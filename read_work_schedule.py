#!/usr/bin/env python
#-*- coding: utf-8 -*-

import sys
import numpy
import dateutil.parser
import openpyxl

from inspect import currentframe, getframeinfo

from errors import log_message, log_error_message, get_stack_trace

# TODO replace with values determined by the defined locations
#max_shift = 10
# max_location = 5
#max_day = 5

# TODO replace with values extracted from the Excel spreadsheet
# Sector and direction meetings:
# - day (0 = Monday)
# - start slot (0 = 8:00)
# - end slot (0 = 8:00)
"""
meeting_slots = {
    'dir': (3, 0, 5),
    'spi': (1, 5, 6),
    'cado': (1, 1, 2),
    'search': (1, 3, 4)
}
"""

# IDEA 2022-02-25 what if we used non-integer hours? Start at 8:30 for example?

def read_work_schedules(xlsx_filename, log_output, error_output):
    log_message(log_output, 'Input file: ' + xlsx_filename)
    wb_obj = openpyxl.load_workbook(xlsx_filename)

    sheet = wb_obj['jours']
    weekdays = {}
    max_day = None
    for row in sheet.iter_rows():
        if len(row) > 0:
            if row[0] is not None:
                cells = [cell.value for cell in row]
                if cells[0] is not None:
                    try:
                        number = int(cells[0])
                    except ValueError:
                        pass
                    value = cells[1]
                    weekdays[number] = value
    max_day = number + 1
    log_message(log_output, f'{max_day} days, namely:  {weekdays}')

    # Definition of daily shifts
    sheet = wb_obj['shifts']
    shifts = []
    for row in sheet.iter_rows():
        if len(row) > 0:
            if row[0] is not None:
                cells = [cell.value for cell in row]
                if cells[0] is not None:
                    parsed_cell = cells[0].replace('h', ':').split(':')
                    time = int(parsed_cell[0]) * 60
                    try:
                        time += int(parsed_cell[1])
                    except (IndexError, ValueError) as e:
                        pass
                    try:
                        parsed_cell = cells[1].replace('h', ':').split(':')
                        value = int(parsed_cell[0]) * 60
                        try:
                            value += int(parsed_cell[1])
                        except (IndexError, ValueError) as e:
                            pass
                    except ValueError:
                        value = 0
                    # log_message(log_output, cells, time, value)
                    shifts.append((time, value))
    log_message(log_output, 'Shift definitions: ' + str(shifts))
    max_shift = len(shifts)
    shift_starts = [x[0] for x in shifts]

    # Get locations from relevant sheet
    sheet = wb_obj['guichets']
    locations = {}
    for row in sheet.iter_rows():
        if len(row) > 0:
            cells = [cell.value for cell in row]
            log_message(log_output, f'guichets tab cells: {cells}')
            if cells[0] is not None:
                name = cells[1]
                locations[cells[0]] = {'name': cells[1], 'times': {}}
                day = -1
                for x in cells[2:2+len(weekdays.keys())]:
                    day += 1
                    times = []
                    if x is None or not x[0].isdigit():
                        log_message(log_output, f'{x} is not a time')
                    else:
                        location_hours = x.split('-')
                        for l in location_hours:
                            hh = int(l.lower().split('h')[0])*60
                            mm = l.lower().split('h')[1]
                            if len(mm) > 0:
                                mm = int(mm)
                            else:
                                mm = 0
                            times.append(hh)
                        #log_message(log_output, 'stuff: ', times, shifts, [x[0] for x in shifts if x[0] <= times[0]])
                        #log_message(log_output, [(x[0], x[0] + x[1], times[1]) for x in shifts ])
                        absolute_times = (max([x[0] for x in shifts if x[0] <= times[0]]),
                                          min([x[0] for x in shifts if x[0] + x[1] >= times[1]]))
                    log_message(log_output, 'Absolute times (minutes): ' + str(absolute_times))
                    locations[cells[0]]['times'][day] = {'start': shift_starts.index(absolute_times[0]),
                        'end': shift_starts.index(absolute_times[1])}

    sheet = wb_obj['quotas']
    quota = {}
    for row in sheet.iter_rows():
        if len(row) > 0:
            cells = [cell.value for cell in row]
            log_message(log_output, f'{cells}')
            if cells[0] is not None:
                cells = [cell.value for cell in row]
                category = cells[0]
                # 2022-05-23 scale up or down if we deal with periods different from 1 business week
                worked = int(cells[1])
                reserve = int(cells[2])
                max_days = int(cells[3])
                quota[category] = (worked, reserve, max_days)

    log_message(log_output, str(locations))
    max_location = len(locations.keys())
    log_message(log_output, f'{max_location} {max_shift}')

    sheet = wb_obj['séances']
    meeting_slots = {}
    # TEST does this work with 2h shifts?
    # We want the first and last **unavailable** slots
    for row in sheet.iter_rows():
        if len(row) > 0:
            if row[0] is not None:
                cells = [cell.value for cell in row]
                if cells[0] is not None:
                    group = cells[0]
                    day = cells[1]     
                    times = []
                    ktimes = -1               
                    for x in cells[2:4]:
                        #log_message(log_output, cells[2:4], cells[2], cells[3], x)
                        ktimes += 1
                        if x is None or not x[0].isdigit():
                            log_message(log_output, f'{x} is not a time')
                        else:
                            # log_message(log_output, [x, x.lower().split('h')])
                            hh = int(x.lower().split('h')[0])*60
                            minutes = x.lower().split('h')[1]
                            if len(minutes) == 0:
                                minutes = '0'
                            mm = int(minutes)
                            log_message(log_output, f'{group} meeting: {hh + mm}, {[x[0] for x in shifts if x[0] <= hh + mm]}')
                            if len([x[0] for x in shifts if x[0] <= hh + mm]) > 0:
                                times.append(max([x[0] for x in shifts if x[0] <= hh + mm]))
                                times.append(min([x[0] for x in shifts if x[0] >= hh + mm]))
                            else:
                                #log_message(log_output, min(shifts))
                                times.append(min([x[0] for x in shifts]))
                                times.append(min([x[0] for x in shifts]))
                            if times[-1] > shifts[-1][0]:
                                pass
                    log_message(log_output, f'{group} meeting times on day {day}: {times}')
                    log_message(log_output, f'{group} meeting shifts on day {day}: {shifts}')

                    meeting_slots[group] = (day, [x[0] for x in shifts].index(times[0]), [x[0] for x in shifts].index(times[3]) -1)

    sheet = wb_obj['guichetiers']
    n = -1
    availability = []
    librarians = {}
    for row in sheet.iter_rows():
        if len(row) > 0:
            if row[0] is not None:
                cells = [cell.value for cell in row]
                if cells[0] is not None:
                    name = cells[0].strip()
                    log_message(log_output, name)
                    n += 1
                    librarians[n] = {'name': name}
                    librarians[n]['sector'] = cells[max_day+2].strip()
                    # we must accept integers and convert to string if necessary
                    librarians[n]['type'] = f'{cells[max_day+3]}'.strip()
                    librarians[n]['prefered_length'] = cells[max_day+4]
                    # TODO replace fixed constants!
                    new_roster = numpy.zeros(shape=(max_day, max_shift, max_location), dtype=numpy.int8)
                    d = -1
                    # Extract extended work hours; skip comment columns
                    for x in cells[1:1+max_day]:
                        # TODO probably not valid for 2h shifts
                        # new day
                        d += 1
                        if x is None or not x[0].isdigit():
                            log_message(log_output, f'{x} is not a time')
                        else:
                            if not x[-1].isdigit():
                                x += '00'
                            x = x.lower().replace('/', '-').replace("–", "-").replace(';','-')
                            x = x.replace('.', 'h').replace(':', 'h').replace('hh', 'h').replace('h-', 'h00-')
                            boundaries = [b.strip() for b in x.split('-')]
                            if len(boundaries) < 2:
                                # something unexpected was found, boundaries were not read properly
                                log_message(error_output, f'WTF {x}')
                                log_error_message(error_output, f'WTF {x}')
                            else:
                                log_message(log_output, f'Boundaries: {boundaries}')
                                int_boundaries = []
                                # rules:
                                # - start time: if minutes are non-zero, round up
                                # - end time: if minutes are non-zero, round down
                                # also: remove 1 from end time to get the beginning of the last slot
                                k = 0
                                for b in boundaries:
                                    split_b = b.strip().split('h')
                                    exact_time = int(split_b[0]) * 60
                                    try:
                                        exact_time += int(split_b[1])
                                    except (IndexError, ValueError) as e:
                                        pass

                                    int_boundaries.append(exact_time)
                                    k += 1
                                #log_message(log_output, 'int_boundaries: ', int_boundaries)
                                for slot in range(len(int_boundaries) // 2):
                                    # for special vacation times, a possible staffmember slot could en before
                                    # the shifts begin...
                                    if int_boundaries[slot*2+1] <= min([x[0] for x in shifts]):
                                        continue
                                    # Or the end of the last shift could take place before  slot begins
                                    if int_boundaries[slot*2] >= max([x[0] for x in shifts]):
                                        continue

                                    if shifts[0][0] <= int_boundaries[slot*2]:
                                        lower_time = min([x[0] for x in shifts if x[0] >= int_boundaries[slot*2]])
                                    else:
                                        lower_time = shifts[0][0]
                                    lower_slot = [x[0] for x in shifts].index(lower_time)

                                    # log_message(log_output, int_boundaries[slot*2+1],  [x[0] for x in shifts])
                                    upper_time = max([x[0] for x in shifts if x[0] <= int_boundaries[slot*2+1]])
                                    # the last 100% possible slot is one below
                                    upper_slot = [x[0] for x in shifts].index(upper_time) - 1
                                    # unless we are beyond the end of the last shiift
                                    if int_boundaries[slot*2+1] >= shifts[-1][0] + shifts[-1][1]:
                                        upper_slot += 1

                                    #log_message(log_output, f'lower-upper: {lower_time}-{upper_time}')
                                    log_message(log_output, f'lower-upper: ({lower_slot})-({upper_slot})')
                                    k = lower_slot
                                    while k <= upper_slot:
                                        # log_message(log_output, [d, k])
                                        for lo in range(max_location):
                                            new_roster[d][k][lo] = 1
                                        k += 1
            
                    availability.append(new_roster)
        else:
            break
    log_message(log_output, '')

    sheet = wb_obj['règles']
    rules = {}
    for row in sheet.iter_rows():
        if len(row) > 0:
            if row[0] is not None:
                cells = [cell.value for cell in row]
                if cells[0] is not None:
                    name = cells[0]
                    try:
                        value = int(cells[1])
                    except ValueError:
                        value = 0
                    rules[name] = (value > 0)
    
    return availability, librarians, locations, quota, meeting_slots, rules, weekdays, shifts


def check_minima(log_output, error_output, availabilities, librarians, locations, quota, meeting_slots, rules, weekdays, shifts):
    """
    Perform some minimal checks: are there slots where we just don't have any available people to begwin with?
    """
    max_shift = len(shifts)
    max_location = len(locations)
    max_day = len(weekdays)
    test_roster = numpy.zeros(shape=(max_day, max_shift, max_location), dtype=numpy.int8)
    minimal_roster = numpy.zeros(shape=(max_day, max_shift, max_location), dtype=numpy.int8)

    msg = ''
    for d in range(max_day):
        for s in range(max_shift):
            for l in range(max_location):
                for n in range(len(librarians)):
                    test_roster[d][s][l] += availabilities[n][d][s][l]
            frameinfo = getframeinfo(currentframe())
            log_message(log_output, f'({frameinfo.filename}:{frameinfo.lineno + 1}) d {d} ({weekdays[d]}) s {s} l {l}: roster = {test_roster[d][s][l]}')
            if test_roster[d][s][0] < sum([shifts[s][0] - shifts[s][0]  >= locations[l]['times'][d]['start'] \
                and shifts[-1][0] - shifts[s][0] <= locations[l]['times'][d]['start'] for l in range(max_location)]):
                    hh = '{:0>2}'.format(shifts[s][0]//60)
                    mm = '{:0>2}'.format(shifts[s][0]%60)
                    msg += f'Warning: not enough staff to fill the {weekdays[d]} {hh}h{mm} slot<br/>\n'

    for d in range(max_day):
        log_message(log_output, weekdays[d])
        for s in range(max_shift):
            frameinfo = getframeinfo(currentframe())
            log_message(log_output, f'({frameinfo.filename}:{frameinfo.lineno + 1}) {[test_roster[d][s][0] for l in range(max_location)]}')
    frameinfo = getframeinfo(currentframe())
    log_message(log_output, f'({frameinfo.filename}:{frameinfo.lineno + 1}) msg "{msg}"')

    return msg


if __name__ == '__main__':
    if len(sys.argv) < 2:
        log_message(log_output, f'Syntaxe: {sys.argv[0]} <XLSX filename>')
        log_message(log_output, 'Le fichier XLSX doit contenir les horaires étendus des collaborateurs')
        exit(1)
    else:
        availabilities, librarians, locations, quota, meeting_slots, rules, weekdays, shifts = read_work_schedules(sys.argv[1])
        msg = check_minima(availabilities, librarians, locations, quota, meeting_slots, rules, weekdays, shifts)
        outfile = open('work_schedule.py', 'w')
        outfile.write('from numpy import array, int8\n\n')
        outfile.write(f'librarians = {str(librarians)}\n')
        outfile.write(f'shift_requests = {str(availabilities)}\n')
        outfile.write(f'\nmeeting_slots = {str(meeting_slots)}\n')
        outfile.write(f'\nlocations = {str(locations)}\n')
        outfile.write(f'\nquota = {str(quota)}\n')
        outfile.write(f'\nrules = {str(rules)}\n')
        outfile.write(f'\nweekdays = {str(weekdays)}\n')
        outfile.write(f'\ndesk_shifts = {str(shifts)}\n')
        msg = msg.replace("\n", "\\n")
        outfile.write(f'\nmsg = "{msg}"\n')
